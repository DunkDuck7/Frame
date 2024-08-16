import os
import json
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.styles import numbers

# --------------------------------------------------------------------------------------------------------
file_path = 'DataFrame.xlsx'

def json_to_python(file_name):

        # LOAD JSON FILE INTO A PYTHON DATA AND STORED IN A VARIABLE
        with open(file_name, 'r', encoding='utf-8') as file:
                json_data = json.load(file)

        # DECLARING ARRAYS, VARIABLES AND DICTIONARY
        records = json_data['data']['records']
        num_records = len(records)

        # --------------------------------------------------------------------------------------------------------

        ExcelFormat = {}

        running_num = []
        branch_code = []
        contract_no = []
        coustomer_fulln = []
        sum_assured = []
        insurer_code = []
        net_premium = []
        total_premium = []
        latest_payment = []
        sales_premium = []
        received_premium = []
        discount = []
        wht = []
        variances = []
        coverage_effective_date = []
        coverage_expired_date = []
        chassis_number = []
        car_brand = []
        vehicle_license = []
        province_name = []
        app_no = []
        sales_code = []
        sales_name = []
        loan_product_code = []
        product_category = []
        payment_method = []
        payment_status = []
        activity_status = []
        transaction_status = []
        sales_channel = []
        system_source = []
        transaction_date = []

        # --------------------------------------------------------------------------------------------------------

        # FUNCTION TO CHANGE THE FORMAT OF THE DATA INTO DATE
        def Date_Format(key):    
                item = json_data['data']['records'][x]
                date_string = (item[key])
                date_part = date_string[:10]
                
                day = date_part[-2:]
                month = date_part[5:7]
                year = str(int(date_part[:4]) + 543)

                date_arranged = f"{day}/{month}/{year}"
                return date_arranged

        # EXTRACTS ALL DATA FROM JSON FILE INTO AN ARRAY
        for x in range(num_records):
                running_num.append(x+1)
                item = json_data['data']['records'][x]
                branch_code.append(item['branch']['code'])
                contract_no.append(item['contract_no'])
                coustomer_fulln.append(item['customer_full_name'])
                insurer_code.append(item['insurer']['key'])
                sum_assured.append(item['sum_assured'])
                net_premium.append(item['net_premium'])
                total_premium.append(item['total_premium'])
                Date = (Date_Format('latest_payment_date'))
                latest_payment.append(Date)
                sales_premium.append(item['sales_premium'])
                received_premium.append(item['received_premium'])
                discount.append(item['discount'])
                wht.append(item['wht'])
                variances.append(item['variances'])
                Date = (Date_Format('coverage_effective_date'))
                coverage_effective_date.append(Date)
                Date = (Date_Format('coverage_expired_date'))
                coverage_expired_date.append(Date)
                chassis_number.append(item['chassis_number'])
                car_brand.append(item['car_brand']['brand_name']['en'])
                vehicle_license.append(item['vehicle_license_no'])
                province_name.append(item['province']['name']['th'])
                app_no.append(item['app_no'])
                sales_code.append(item['sales_code'])
                sales_name.append(item['sales_name'])
                loan_product_code.append(item['loan_product']['code'])
                product_category.append(item['product_category']['name']['en'])
                payment_method.append(item['payment_method']['name']['th'])
                payment_status.append(item['payment_status']['name']['en'])
                activity_status.append(item['activity_status']['name']['th'])
                transaction_status.append(item['transaction_status']['name']['th'])
                sales_channel.append(item['sale_channel']['name']['th'])
                system_source.append(item['source_system']['name']['th'])
                Date = (Date_Format('transaction_date'))
                transaction_date.append(Date)

        # UPDATING THE DICTIONARY
        ExcelFormat = {
                "#" : running_num,
                "branch_code" : branch_code,
                "contract_no" : contract_no,
                "coustomer_fulln" : coustomer_fulln,
                "sum_assured" : sum_assured,
                "insurer_code" : insurer_code,
                "net_premium" : net_premium,
                "total_premium" : total_premium ,
                "latest_payment" : latest_payment,
                "sales_premium" : sales_premium,
                "sales_premium" : sales_premium,
                "discount" : discount,
                "wht" : wht,
                "variances" : variances,
                "coverage_effective_date" : coverage_effective_date,
                "coverage_expired_date" : coverage_expired_date,
                "chassis_number" : chassis_number,
                "car_brand" : car_brand,
                "vehicle_license" : vehicle_license,
                "province_name" : province_name,
                "app_no" : app_no,
                "sales_code" : sales_code,
                "sales_name" : sales_name,
                "loan_product_code" : loan_product_code,
                "product_category" : product_category,
                "payment_method" : payment_method,
                "payment_status" : payment_status,
                "activity_status" : activity_status,
                "transaction_status" : transaction_status,
                "sales_channel" : sales_channel,
                "system_source" : system_source,
                "transaction_date" : transaction_date,                  
        }

        return ExcelFormat

def python_to_excel(data):
        global file_path
        #STORING THE DICTIONARY INTO A DATAFRAME
        df = pd.DataFrame(data)    

        # --------------------------------------------------------------------------------------------------------  

        try:

                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        
        # --------------------------------------------------------------------------------------------------------        
                        
                        # FUNCTION TO CONVERT STR TO INT (ALL COLUMN)
                        def convert_columns_to_int(df):
                                for col in df.columns:
                                        # TRY TO CONVERT STR TO INT
                                        try:
                                        # CHECK IF CONVERSION IS POSSIBLE, IF NOT ERROR IS RAISE                                
                                                df[col] = pd.to_numeric(df[col], errors='raise', downcast='integer')
                                        except ValueError: 
                                        # IF CONVERSION FAILS, COLUMN ARE ENSURE AS STRING:
                                                df[col] = df[col].astype(str)

                                return df

                        # CONVERT ALL COLUMNS CHECKED
                        df = convert_columns_to_int(df)

        # --------------------------------------------------------------------------------------------------------

                        # ATTEMPT TO WRITE DATAFRAME TO EXCEL
                        df.to_excel(writer, index=False, sheet_name='Information')
                        print(f"Data successfully written to {file_path}")
        
                        workbook = writer.book
                        sheet = writer.sheets['Information']

        # --------------------------------------------------------------------------------------------------------

                        # FUNCTION APPLY COMMA TO NUMBERS
                        def apply_comma_format_to_integers(sheet, df):
                                # ITERATE THROUGH ALL COLUMNS
                                for col_index in range(1, len(df.columns) + 1):
                                        col_letter = get_column_letter(col_index)
                                        col_name = df.columns[col_index - 1]
                        
                                        # CHECK IF COLUMN IS INTEGER TYPE
                                        if pd.api.types.is_any_real_numeric_dtype(df[col_name]):
                                                for row in range(2, sheet.max_row + 1):  # SKIP HEADER ROW
                                                        cell = sheet[f'{col_letter}{row}']
                                                        if isinstance(cell.value, (int, float)):  # ONLY APPLY TO NUMBERS
                                                                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                        
        # --------------------------------------------------------------------------------------------------------

                        # CONVERT COLUMN['branch_code'] INTO STR
                        df['branch_code'] = df['branch_code'].astype(int)
                        df['#'] = df['#'].astype(int)

        # --------------------------------------------------------------------------------------------------------

                        # APPLY COMMA FORMAT
                        apply_comma_format_to_integers(sheet, df)

        # --------------------------------------------------------------------------------------------------------

                        # ADJUST COLUMN WIDTH
                        def adjust_cell():
                                for col in sheet.columns:
                                        max_length = 0
                                        column = col[0].column_letter
                                        for cell in col:
                                                try:
                                                        max_length = max(max_length, len(str(cell.value)))
                                                except:
                                                        pass
                                                adjusted_width = (max_length + 4)
                                                sheet.column_dimensions[column].width = adjusted_width
                                
                                # ADJUST ROW HEIGHT
                                sheet.row_dimensions[1].height = 15

        # --------------------------------------------------------------------------------------------------------

                        # DEFINE BORDER
                        thin_border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                        )

                        # APPLYING BORDER TO ALL CELLS
                        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                                for cell in row:
                                        cell.border = thin_border

        # --------------------------------------------------------------------------------------------------------

                        # DEFINE FILL STYLES
                        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
                        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

                        # APPLYING THE CONDITION TO A COLUMN   
                        col_index = df.columns.get_loc('variances') + 1  # Get the 1-based index of 'Column1'
                        col_letter = get_column_letter(col_index)

                        # HIGHLIGHT CELL BASED ON VALUE
                        # APPLY GREEN IF <0
                        sheet.conditional_formatting.add(
                                f'{col_letter}2:{col_letter}{sheet.max_row}',
                                FormulaRule(formula=[f'{col_letter}2>0'], fill=green_fill)
                        )

                        # APPLY RED IF >0
                        sheet.conditional_formatting.add(
                                f'{col_letter}2:{col_letter}{sheet.max_row}',
                                FormulaRule(formula=[f'{col_letter}2<0'], fill=red_fill)
                        )

        # --------------------------------------------------------------------------------------------------------

                        # SUM OF THE COLUMN['sales_premium']
                        Sum_column = df['sales_premium'].sum()
                        Summary_Cell = 'J23'
                        # ADD SUM INTO THE SUMMARY_CELL 
                        sheet[Summary_Cell] = Sum_column
                        cell = sheet[f'{Summary_Cell}']
                        # FORMAT 'Summary_Cell' TO NUMBER COMMA
                        cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        # --------------------------------------------------------------------------------------------------------

                        # DEFINE DOUBLE UNDERLINE
                        double_underline = Border(
                        bottom=Side(style='double')
                        )

                        # APPLY DOUBLE UNDERLINE TO SUMMARY_CELL
                        sheet[Summary_Cell].border = double_underline
                        # ADD 'Summary' ONTO CELL
                        sheet['I23'] = "Summary"

        # --------------------------------------------------------------------------------------------------------

                        # DEFINE DIAGONAL BORDER STYLE
                        diagonal_border = Border(
                                diagonal=Side(style="thick", color="000000"),  # Diagonal line style (thin) and color (black)
                                diagonalUp=True,   # Diagonal line from bottom-left to top-right
                                diagonalDown=True  # Diagonal line from top-left to bottom-right
                        )

                        # ITERATE THROUGH EVERY ROW_CELL AND COLUMN_CELL
                        for col_index in range(1, len(df.columns) + 1):
                                        col_letter = get_column_letter(col_index)
                                        col_name = df.columns[col_index - 1]
                                        for row in range(2, sheet.max_row + 1):
                                                if cell.value == "":    
                                                        cell.border = diagonal_border

                        # APPLY CONDITION          
                        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                                for cell in row:
                                        if cell.value == "":    
                                                cell.border = diagonal_border

        # --------------------------------------------------------------------------------------------------------
                        
                        adjust_cell()

                        # SAVE WORKBOOK AFTER FORMATTING
                        workbook.save(file_path)

                return df

        # --------------------------------------------------------------------------------------------------------

        except PermissionError as pe:
                # Handle the case where the file is open or has restricted permissions
                print(f"{pe}: Please close the file destination")  

        except Exception as e:
                # Handle any other exceptions that might occur
                print(f"An unexpected error occurred: {e}")

        # --------------------------------------------------------------------------------------------------------

